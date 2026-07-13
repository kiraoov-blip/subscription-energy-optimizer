from __future__ import annotations

import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence, Tuple, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from ortools.sat.python import cp_model

SEASONS = ("봄가을", "여름", "겨울")
DAY_TYPES = ("주중", "주말")
MODES = ("편의 우선", "균형", "절약 우선")
CURRENT_RATE_LABEL = "현행 한계단가"

MODE_COLUMN = {
    "편의 우선": "최대감축률_편의",
    "균형": "최대감축률_균형",
    "절약 우선": "최대감축률_절약",
}

# CP-SAT objective coefficients. The values are relative scores, not retail tariffs.
SYSTEM_WEIGHT_SCALE = 20
PEAK_PENALTY_PER_WH = 100
ENERGY_SHIFT_PENALTY_PER_WH = 2
CYCLE_DELAY_PENALTY_SCALE = 2
HVAC_DISCOMFORT_PER_WH = {"편의 우선": 500, "균형": 300, "절약 우선": 180}
BEHAVIOR_DISCOMFORT_PER_WH = {"편의 우선": 800, "균형": 500, "절약 우선": 300}


@dataclass
class WorkbookData:
    path: str
    appliances: List[str]
    meta: pd.DataFrame
    controls: pd.DataFrame
    plans: pd.DataFrame
    settings: Dict[str, object]
    weights: pd.DataFrame
    scenarios: Dict[Tuple[str, str], pd.DataFrame]


@dataclass
class SolveResult:
    summary: Dict[str, object]
    hourly: pd.DataFrame
    appliances: pd.DataFrame


def _num(value, default: float = 0.0) -> float:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _yes(value: object) -> bool:
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1", "예", "사용", "적용"}


def parse_hour_window(text: object) -> List[int]:
    """Parse strings such as '21-7' or '9-18,20-7' into hour integers."""
    if text is None or (isinstance(text, float) and math.isnan(text)):
        return list(range(24))
    s = str(text).strip().replace("~", "-").replace("시", "")
    if not s:
        return list(range(24))
    hours: set[int] = set()
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" not in part:
            try:
                hours.add(int(float(part)) % 24)
            except ValueError:
                continue
            continue
        a_text, b_text = part.split("-", 1)
        try:
            a = int(float(a_text)) % 24
            b_raw = int(float(b_text))
            b = b_raw % 24
        except ValueError:
            continue
        if a == b and b_raw not in {0, 24}:
            hours.update(range(24))
        elif a < b and b_raw <= 24:
            hours.update(range(a, b))
        else:
            hours.update(range(a, 24))
            hours.update(range(0, b))
    return sorted(hours) if hours else list(range(24))


def _hours_to_text(values_wh: Sequence[int]) -> str:
    active = [str(h) for h, v in enumerate(values_wh) if v > 0]
    return ",".join(active) if active else "-"


def load_workbook_data(path: Union[str, os.PathLike]) -> WorkbookData:
    """Load the expanded input workbook. Falls back to safe defaults if config sheets are absent."""
    path = str(path)
    if not Path(path).exists():
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")

    meta = pd.read_excel(path, sheet_name="가전_입력가정")
    meta = meta.dropna(subset=["가전기기"]).copy()
    appliances = [str(x).strip() for x in meta["가전기기"].tolist()]

    xls = pd.ExcelFile(path)
    scenarios: Dict[Tuple[str, str], pd.DataFrame] = {}
    for season in SEASONS:
        for day in DAY_TYPES:
            sheet = f"{season}_{day}"
            if sheet not in xls.sheet_names:
                raise ValueError(f"필수 시트가 없습니다: {sheet}")
            df = pd.read_excel(path, sheet_name=sheet, nrows=24)
            missing = [a for a in appliances if a not in df.columns]
            if missing:
                raise ValueError(f"{sheet} 시트에 가전 열이 없습니다: {missing[:5]}")
            for col in appliances:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            df["시간"] = pd.to_numeric(df["시간"], errors="coerce").fillna(pd.Series(np.arange(len(df)), index=df.index)).astype(int)
            scenarios[(season, day)] = df[["시간"] + appliances].copy()

    wb = load_workbook(path, data_only=True, read_only=False)

    # Plans and general settings
    plans_rows: List[Dict[str, object]] = []
    settings: Dict[str, object] = {
        "월 주중일수": 22,
        "월 주말일수": 8,
        "초과단가_저(원/kWh)": 200,
        "초과단가_중(원/kWh)": 300,
        "초과단가_고(원/kWh)": 400,
        "현행한계단가_봄가을(원/kWh)": 307.3,
        "현행한계단가_여름(원/kWh)": 242.3,
        "현행한계단가_겨울(원/kWh)": 307.3,
        "기후환경요금단가(원/kWh)": 0.0,
        "연료비조정단가(원/kWh)": 0.0,
        "부가가치세율": 0.10,
        "전력산업기반기금율": 0.027,
        "상세결과_계절": "여름",
        "상세결과_요금제": "기본형",
        "상세결과_냉난방모드": "균형",
        "상세결과_초과단가": 300,
    }
    if "요금제_설정" in wb.sheetnames:
        ws = wb["요금제_설정"]
        for r in range(4, 7):
            name = ws.cell(r, 1).value
            if name:
                plans_rows.append({
                    "요금제": str(name),
                    "월 구독료(원)": _num(ws.cell(r, 2).value),
                    "기본 제공량(kWh)": _num(ws.cell(r, 3).value),
                    "주요 대상": ws.cell(r, 4).value,
                    "분석 활성화": ws.cell(r, 5).value,
                })
        for r in range(10, ws.max_row + 1):
            key = ws.cell(r, 1).value
            if key:
                settings[str(key)] = ws.cell(r, 2).value
    if not plans_rows:
        plans_rows = [
            {"요금제": "알뜰형", "월 구독료(원)": 19900, "기본 제공량(kWh)": 200, "주요 대상": "1인 가구", "분석 활성화": "N"},
            {"요금제": "기본형", "월 구독료(원)": 84900, "기본 제공량(kWh)": 450, "주요 대상": "4인 가구", "분석 활성화": "Y"},
            {"요금제": "프리미엄형", "월 구독료(원)": 249900, "기본 제공량(kWh)": 1000, "주요 대상": "고사용량 4인 가구", "분석 활성화": "Y"},
        ]
    plans = pd.DataFrame(plans_rows)

    # Relative time weights
    if "시간대가중치" in wb.sheetnames:
        ws = wb["시간대가중치"]
        weight_rows = []
        for r in range(4, 28):
            h = ws.cell(r, 1).value
            if h is None:
                continue
            weight_rows.append({
                "시간": int(h),
                "봄가을": int(_num(ws.cell(r, 2).value, 2)),
                "여름": int(_num(ws.cell(r, 3).value, 2)),
                "겨울": int(_num(ws.cell(r, 4).value, 2)),
            })
        weights = pd.DataFrame(weight_rows)
    else:
        weights = pd.DataFrame({
            "시간": range(24),
            "봄가을": [1 if h >= 23 or h < 8 else (4 if 17 <= h < 21 else 2) for h in range(24)],
            "여름": [1 if h >= 23 or h < 8 else (4 if 14 <= h < 21 else 2) for h in range(24)],
            "겨울": [1 if h >= 23 or h < 7 else (4 if 8 <= h < 11 or 17 <= h < 21 else 2) for h in range(24)],
        })

    # Appliance controls
    control_rows: List[Dict[str, object]] = []
    if "가전제어_설정" in wb.sheetnames:
        ws = wb["가전제어_설정"]
        headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        for r in range(4, ws.max_row + 1):
            app = ws.cell(r, 2).value
            if not app:
                continue
            row = {str(headers[c - 1]): ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
            control_rows.append(row)
    controls = pd.DataFrame(control_rows)
    if controls.empty:
        controls = pd.DataFrame({
            "가전기기": appliances,
            "제어유형": ["비제어"] * len(appliances),
            "적용여부": ["N"] * len(appliances),
        })
    controls["가전기기"] = controls["가전기기"].astype(str).str.strip()
    controls = controls.set_index("가전기기", drop=False)

    wb.close()
    return WorkbookData(path, appliances, meta, controls, plans, settings, weights, scenarios)


def get_overage_rates(data: WorkbookData, season: str) -> List[Tuple[str, float]]:
    s = data.settings
    current = _num(s.get(f"현행한계단가_{season}(원/kWh)"), 307.3 if season != "여름" else 242.3)
    return [
        ("200원", _num(s.get("초과단가_저(원/kWh)"), 200)),
        ("300원", _num(s.get("초과단가_중(원/kWh)"), 300)),
        ("400원", _num(s.get("초과단가_고(원/kWh)"), 400)),
        (CURRENT_RATE_LABEL, current),
    ]


def current_residential_bill(
    kwh: float,
    season: str,
    climate_rate: float = 0.0,
    fuel_rate: float = 0.0,
    vat_rate: float = 0.10,
    fund_rate: float = 0.027,
) -> float:
    """Approximate 2026-06-01 residential low-voltage bill.

    Climate-environment and fuel-adjustment rates are user inputs because they can change.
    """
    kwh = max(float(kwh), 0.0)
    if season == "여름":
        basic = 730 if kwh <= 300 else (1260 if kwh <= 450 else 6060)
        tiers = [(300, 105.0), (150, 174.0), (550, 242.3)]
    else:
        basic = 910 if kwh <= 200 else (1600 if kwh <= 400 else 7300)
        tiers = [(200, 120.0), (200, 214.6), (600, 307.3)]

    remaining = kwh
    energy_charge = 0.0
    for qty, rate in tiers:
        used = min(remaining, qty)
        energy_charge += used * rate
        remaining -= used
        if remaining <= 1e-9:
            break
    if remaining > 0:
        # Super-user tariff applies in summer and winter above 1,000 kWh.
        super_rate = 736.2 if season in {"여름", "겨울"} else tiers[-1][1]
        energy_charge += remaining * super_rate

    electricity = basic + energy_charge + kwh * (climate_rate + fuel_rate)
    vat = round(electricity * vat_rate)
    fund = math.floor(electricity * fund_rate / 10.0) * 10.0
    return electricity + vat + fund


def _control_row(data: WorkbookData, appliance: str) -> Mapping[str, object]:
    if appliance in data.controls.index:
        row = data.controls.loc[appliance]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]
        return row.to_dict()
    return {"가전기기": appliance, "제어유형": "비제어", "적용여부": "N"}


def _meta_average_kw(data: WorkbookData, appliance: str) -> float:
    hit = data.meta.loc[data.meta["가전기기"].astype(str).str.strip() == appliance]
    if hit.empty:
        return 0.0
    return _num(hit.iloc[0].get("모델링 활성 평균(kW)"), 0.0)


def solve_month(
    data: WorkbookData,
    season: str,
    plan_name: str,
    mode: str,
    overage_label: str,
    overage_rate: float,
    customer_override: bool = False,
    max_solve_seconds: float = 10.0,
) -> SolveResult:
    if season not in SEASONS:
        raise ValueError(f"지원하지 않는 계절: {season}")
    if mode not in MODES:
        raise ValueError(f"지원하지 않는 냉난방 모드: {mode}")
    plan_hit = data.plans.loc[data.plans["요금제"] == plan_name]
    if plan_hit.empty:
        raise ValueError(f"요금제를 찾을 수 없습니다: {plan_name}")
    plan = plan_hit.iloc[0]
    plan_fee = _num(plan["월 구독료(원)"])
    plan_limit_kwh = _num(plan["기본 제공량(kWh)"])

    weekday_days = int(_num(data.settings.get("월 주중일수"), 22))
    weekend_days = int(_num(data.settings.get("월 주말일수"), 8))
    day_counts = {"주중": weekday_days, "주말": weekend_days}
    weight_by_hour = {int(r["시간"]): int(r[season]) for _, r in data.weights.iterrows()}

    model = cp_model.CpModel()
    objective_terms: List[cp_model.LinearExpr] = []
    hourly_load_vars: Dict[Tuple[str, int], cp_model.IntVar] = {}
    app_exprs: Dict[Tuple[str, str, int], Union[int, cp_model.IntVar, cp_model.LinearExpr]] = {}
    base_wh: Dict[Tuple[str, str], List[int]] = {}
    app_aux: Dict[Tuple[str, str], Dict[str, object]] = {}

    for day in DAY_TYPES:
        day_df = data.scenarios[(season, day)]
        count = day_counts[day]
        for app in data.appliances:
            values = [max(0, int(round(v * 1000))) for v in day_df[app].tolist()]
            base_wh[(day, app)] = values
            cfg = _control_row(data, app)
            ctype = str(cfg.get("제어유형", "비제어")).strip()
            enabled = _yes(cfg.get("적용여부", "N")) and not customer_override
            aux: Dict[str, object] = {"type": ctype, "enabled": enabled}
            app_aux[(day, app)] = aux

            if not enabled or ctype in {"비제어", "알림", ""} or sum(values) == 0:
                for h in range(24):
                    app_exprs[(day, app, h)] = values[h]
                continue

            if ctype == "에너지이동":
                allowed_text = cfg.get("주중 허용시간" if day == "주중" else "주말 허용시간", "")
                allowed = set(parse_hour_window(allowed_text))
                total = sum(values)
                avg_wh = int(round(_meta_average_kw(data, app) * 1000))
                min_capacity = math.ceil(total / max(len(allowed), 1))
                max_per_hour = max(max(values), avg_wh, min_capacity, 1)
                vars_h: List[cp_model.IntVar] = []
                diffs: List[cp_model.IntVar] = []
                for h in range(24):
                    ub = max_per_hour if h in allowed else 0
                    v = model.NewIntVar(0, ub, f"move_{day}_{app}_{h}")
                    vars_h.append(v)
                    app_exprs[(day, app, h)] = v
                    d = model.NewIntVar(0, max(max_per_hour, values[h]), f"diff_{day}_{app}_{h}")
                    model.AddAbsEquality(d, v - values[h])
                    diffs.append(d)
                model.Add(sum(vars_h) == total)
                priority = max(1, int(_num(cfg.get("지연불편가중치"), 1)))
                objective_terms.extend(d * ENERGY_SHIFT_PENALTY_PER_WH * priority * count for d in diffs)
                aux.update({"vars": vars_h, "diffs": diffs})
                continue

            if ctype == "사이클이동":
                nonzero = [h for h, v in enumerate(values) if v > 0]
                if not nonzero:
                    for h in range(24):
                        app_exprs[(day, app, h)] = values[h]
                    continue
                original_start = nonzero[0]
                profile = values[original_start : nonzero[-1] + 1]
                allowed_text = cfg.get("주중 허용시간" if day == "주중" else "주말 허용시간", "")
                allowed = set(parse_hour_window(allowed_text))
                candidates: List[int] = []
                for s in range(24):
                    used_hours = {(s + i) % 24 for i, p in enumerate(profile) if p > 0}
                    if used_hours.issubset(allowed):
                        candidates.append(s)
                if original_start not in candidates:
                    candidates.append(original_start)
                candidates = sorted(set(candidates))
                starts = {s: model.NewBoolVar(f"start_{day}_{app}_{s}") for s in candidates}
                model.Add(sum(starts.values()) == 1)
                for h in range(24):
                    terms = []
                    for s, b in starts.items():
                        offset = (h - s) % 24
                        if offset < len(profile) and profile[offset] > 0:
                            terms.append(b * profile[offset])
                    app_exprs[(day, app, h)] = sum(terms) if terms else 0
                priority = max(1, int(_num(cfg.get("지연불편가중치"), 1)))
                total = sum(profile)
                for s, b in starts.items():
                    distance = min((s - original_start) % 24, (original_start - s) % 24)
                    objective_terms.append(b * distance * total * priority * CYCLE_DELAY_PENALTY_SCALE * count)
                aux.update({"starts": starts, "profile": profile, "original_start": original_start})
                continue

            if ctype == "출력조정":
                reduction_pct = _num(cfg.get(MODE_COLUMN[mode]), 0.0)
                reductions: List[Union[int, cp_model.IntVar]] = []
                for h, base in enumerate(values):
                    max_reduction = int(math.floor(base * reduction_pct + 1e-9))
                    if max_reduction <= 0:
                        app_exprs[(day, app, h)] = base
                        reductions.append(0)
                    else:
                        red = model.NewIntVar(0, max_reduction, f"reduce_{day}_{app}_{h}")
                        app_exprs[(day, app, h)] = base - red
                        reductions.append(red)
                        extra_weight = max(1, int(_num(cfg.get("감축불편가중치"), 1)))
                        objective_terms.append(red * HVAC_DISCOMFORT_PER_WH[mode] * extra_weight * count)
                aux.update({"reductions": reductions})
                continue

            if ctype == "행동제어":
                reduction_pct = _num(cfg.get(MODE_COLUMN[mode]), 0.0)
                active_hours = [h for h, v in enumerate(values) if v > 0]
                max_hours_value = cfg.get("최대사용시간_주중" if day == "주중" else "최대사용시간_주말")
                max_hours = int(_num(max_hours_value, len(active_hours))) if active_hours else 0
                min_retain_hours = math.ceil(len(active_hours) * (1.0 - reduction_pct))
                on_vars: Dict[int, cp_model.IntVar] = {}
                for h in range(24):
                    if values[h] <= 0:
                        app_exprs[(day, app, h)] = 0
                        continue
                    on = model.NewBoolVar(f"behavior_{day}_{app}_{h}")
                    on_vars[h] = on
                    app_exprs[(day, app, h)] = on * values[h]
                    off_energy = values[h] * (1 - on)
                    extra_weight = max(1, int(_num(cfg.get("감축불편가중치"), 1)))
                    objective_terms.append(off_energy * BEHAVIOR_DISCOMFORT_PER_WH[mode] * extra_weight * count)
                if on_vars:
                    hard_max = min(max_hours, len(on_vars))
                    hard_min = min(min_retain_hours, hard_max)
                    model.Add(sum(on_vars.values()) <= hard_max)
                    model.Add(sum(on_vars.values()) >= hard_min)
                aux.update({"on_vars": on_vars})
                continue

            # Unknown type -> fixed for safety
            for h in range(24):
                app_exprs[(day, app, h)] = values[h]

        # Hourly total load
        for h in range(24):
            terms = [app_exprs[(day, app, h)] for app in data.appliances]
            upper = max(1, sum(max(base_wh[(day, app)]) for app in data.appliances) * 2)
            load = model.NewIntVar(0, upper, f"load_{day}_{h}")
            model.Add(load == sum(terms))
            hourly_load_vars[(day, h)] = load
            objective_terms.append(load * weight_by_hour[h] * SYSTEM_WEIGHT_SCALE * count)

    monthly_total = model.NewIntVar(0, 10_000_000, "monthly_total_wh")
    model.Add(monthly_total == sum(hourly_load_vars[(day, h)] * day_counts[day] for day in DAY_TYPES for h in range(24)))
    plan_limit_wh = int(round(plan_limit_kwh * 1000))
    overage = model.NewIntVar(0, 10_000_000, "overage_wh")
    model.Add(overage >= monthly_total - plan_limit_wh)
    model.Add(overage >= 0)
    objective_terms.append(overage * int(round(overage_rate)))  # milliwon

    peak = model.NewIntVar(0, 100_000, "peak_wh")
    model.AddMaxEquality(peak, list(hourly_load_vars.values()))
    objective_terms.append(peak * PEAK_PENALTY_PER_WH)

    model.Minimize(sum(objective_terms))
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = max_solve_seconds
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"최적화 해를 찾지 못했습니다. 상태={solver.StatusName(status)}")

    hourly_rows: List[Dict[str, object]] = []
    app_rows: List[Dict[str, object]] = []
    base_month_wh = 0
    opt_month_wh = int(solver.Value(monthly_total))
    base_peak_wh = 0
    opt_peak_wh = int(solver.Value(peak))
    total_shift_wh = 0.0
    total_curtail_wh = 0.0
    weighted_shift_hours = 0.0

    solved_values: Dict[Tuple[str, str], List[int]] = {}
    for day in DAY_TYPES:
        count = day_counts[day]
        for app in data.appliances:
            vals: List[int] = []
            for h in range(24):
                expr = app_exprs[(day, app, h)]
                if isinstance(expr, int):
                    vals.append(expr)
                else:
                    vals.append(int(solver.Value(expr)))
            solved_values[(day, app)] = vals
            base = base_wh[(day, app)]
            base_total = sum(base)
            opt_total = sum(vals)
            shifted = sum(abs(vals[h] - base[h]) for h in range(24)) / 2.0
            curtailed = max(0, base_total - opt_total)
            total_shift_wh += shifted * count
            total_curtail_wh += curtailed * count
            cfg = _control_row(data, app)
            ctype = str(cfg.get("제어유형", "비제어"))
            original_hours = _hours_to_text(base)
            optimized_hours = _hours_to_text(vals)
            if ctype == "사이클이동" and base_total > 0:
                orig_start = next((h for h, v in enumerate(base) if v > 0), 0)
                opt_start = next((h for h, v in enumerate(vals) if v > 0), orig_start)
                weighted_shift_hours += min((opt_start - orig_start) % 24, (orig_start - opt_start) % 24) * count
            app_rows.append({
                "계절": season,
                "요일": day,
                "가전기기": app,
                "제어유형": ctype,
                "적용여부": cfg.get("적용여부", "N"),
                "기준사용량(kWh/일)": base_total / 1000.0,
                "최적사용량(kWh/일)": opt_total / 1000.0,
                "이동량(kWh/일)": shifted / 1000.0,
                "감축량(kWh/일)": curtailed / 1000.0,
                "기준사용시간": original_hours,
                "최적사용시간": optimized_hours,
            })

        for h in range(24):
            base_load = sum(base_wh[(day, app)][h] for app in data.appliances)
            opt_load = int(solver.Value(hourly_load_vars[(day, h)]))
            base_peak_wh = max(base_peak_wh, base_load)
            base_month_wh += base_load * count
            hourly_rows.append({
                "계절": season,
                "요일": day,
                "시간": h,
                "시간대": f"{h:02d}:00~{(h+1)%24:02d}:00",
                "일수반영": count,
                "상대가중치": weight_by_hour[h],
                "기준부하(kWh)": base_load / 1000.0,
                "최적부하(kWh)": opt_load / 1000.0,
                "차이(kWh)": (opt_load - base_load) / 1000.0,
            })

    base_month_kwh = base_month_wh / 1000.0
    opt_month_kwh = opt_month_wh / 1000.0
    overage_kwh = max(0.0, opt_month_kwh - plan_limit_kwh)
    subscription_bill = plan_fee + overage_kwh * overage_rate
    climate = _num(data.settings.get("기후환경요금단가(원/kWh)"), 0)
    fuel = _num(data.settings.get("연료비조정단가(원/kWh)"), 0)
    vat = _num(data.settings.get("부가가치세율"), 0.10)
    fund = _num(data.settings.get("전력산업기반기금율"), 0.027)
    current_base_bill = current_residential_bill(base_month_kwh, season, climate, fuel, vat, fund)
    current_opt_bill = current_residential_bill(opt_month_kwh, season, climate, fuel, vat, fund)

    # Comfort score: transparent heuristic, not a measured customer satisfaction score.
    shift_penalty_points = min(20.0, total_shift_wh / max(base_month_wh, 1) * 40.0)
    curtail_penalty_points = min(45.0, total_curtail_wh / max(base_month_wh, 1) * 180.0)
    delay_penalty_points = min(15.0, weighted_shift_hours / max(weekday_days + weekend_days, 1) * 0.7)
    comfort_score = max(0.0, 100.0 - shift_penalty_points - curtail_penalty_points - delay_penalty_points)

    # Plan recommendation
    basic_hit = data.plans.loc[data.plans["요금제"] == "기본형"]
    premium_hit = data.plans.loc[data.plans["요금제"] == "프리미엄형"]
    recommendation = "현재 요금제 유지"
    reason = ""
    if not basic_hit.empty and not premium_hit.empty:
        basic_fee = _num(basic_hit.iloc[0]["월 구독료(원)"])
        basic_limit = _num(basic_hit.iloc[0]["기본 제공량(kWh)"])
        premium_fee = _num(premium_hit.iloc[0]["월 구독료(원)"])
        basic_equiv = basic_fee + max(0, opt_month_kwh - basic_limit) * overage_rate
        if plan_name == "기본형":
            if basic_equiv >= premium_fee or overage_kwh >= 150 or comfort_score < 60:
                recommendation = "프리미엄형 전환 검토"
                reasons = []
                if basic_equiv >= premium_fee:
                    reasons.append("기본형+초과요금이 프리미엄형 이상")
                if overage_kwh >= 150:
                    reasons.append("예상 초과량 150kWh 이상")
                if comfort_score < 60:
                    reasons.append("고객 편의지수 60점 미만")
                reason = ", ".join(reasons)
        elif plan_name == "프리미엄형":
            if opt_month_kwh <= basic_limit and basic_equiv < premium_fee and comfort_score >= 80:
                recommendation = "기본형 전환 검토"
                reason = "최적사용량이 기본형 제공량 이내이고 예상 납부액이 더 낮음"

    summary = {
        "계절": season,
        "요금제": plan_name,
        "냉난방모드": mode,
        "초과단가시나리오": overage_label,
        "초과단가(원/kWh)": overage_rate,
        "구독료(원)": plan_fee,
        "제공량(kWh)": plan_limit_kwh,
        "기준월사용량(kWh)": base_month_kwh,
        "최적월사용량(kWh)": opt_month_kwh,
        "월감축량(kWh)": base_month_kwh - opt_month_kwh,
        "초과량(kWh)": overage_kwh,
        "구독최종납부액(원)": subscription_bill,
        "현행요금_기준(원)": current_base_bill,
        "현행요금_최적(원)": current_opt_bill,
        "기준피크(kW)": base_peak_wh / 1000.0,
        "최적피크(kW)": opt_peak_wh / 1000.0,
        "피크감축(kW)": (base_peak_wh - opt_peak_wh) / 1000.0,
        "월이동량(kWh)": total_shift_wh / 1000.0,
        "고객편의지수(100점)": comfort_score,
        "요금제권고": recommendation,
        "권고사유": reason,
        "고객수동해제": "Y" if customer_override else "N",
        "해상태": solver.StatusName(status),
    }
    return SolveResult(summary, pd.DataFrame(hourly_rows), pd.DataFrame(app_rows))


def run_batch(
    data: WorkbookData,
    seasons: Sequence[str] = SEASONS,
    plans: Sequence[str] = ("기본형", "프리미엄형"),
    modes: Sequence[str] = MODES,
    include_all_overage_rates: bool = True,
    customer_override: bool = False,
    max_solve_seconds: float = 5.0,
    detail_selector: Optional[Tuple[str, str, str, Union[str, float]]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    summaries: List[Dict[str, object]] = []
    detail_hourly: Optional[pd.DataFrame] = None
    detail_apps: Optional[pd.DataFrame] = None

    if detail_selector is None:
        detail_selector = (
            str(data.settings.get("상세결과_계절", "여름")),
            str(data.settings.get("상세결과_요금제", "기본형")),
            str(data.settings.get("상세결과_냉난방모드", "균형")),
            data.settings.get("상세결과_초과단가", 300),
        )

    for season in seasons:
        rates = get_overage_rates(data, season)
        if not include_all_overage_rates:
            rates = [rates[1]]
        for plan in plans:
            for mode in modes:
                for label, rate in rates:
                    result = solve_month(
                        data, season, plan, mode, label, rate,
                        customer_override=customer_override,
                        max_solve_seconds=max_solve_seconds,
                    )
                    summaries.append(result.summary)
                    ds, dp, dm, dr = detail_selector
                    rate_match = (str(dr) == label) or (isinstance(dr, (int, float)) and abs(float(dr) - rate) < 1e-6)
                    if season == ds and plan == dp and mode == dm and rate_match:
                        detail_hourly = result.hourly.copy()
                        detail_apps = result.appliances.copy()

    summary_df = pd.DataFrame(summaries)
    if detail_hourly is None or detail_apps is None:
        # Guaranteed fallback: first requested combination, 300-won scenario.
        season = seasons[0]
        plan = plans[0]
        mode = modes[0]
        rates = get_overage_rates(data, season)
        label, rate = rates[1]
        result = solve_month(data, season, plan, mode, label, rate, customer_override, max_solve_seconds)
        detail_hourly = result.hourly
        detail_apps = result.appliances
    return summary_df, detail_hourly, detail_apps


def export_results(
    output_path: Union[str, os.PathLike],
    summary_df: pd.DataFrame,
    detail_hourly: pd.DataFrame,
    detail_apps: pd.DataFrame,
    data: WorkbookData,
) -> str:
    output_path = str(output_path)
    # Useful filtered view: only appliances whose schedule or energy changed.
    changed = detail_apps.loc[
        (detail_apps["이동량(kWh/일)"] > 1e-9)
        | (detail_apps["감축량(kWh/일)"] > 1e-9)
        | (detail_apps["기준사용시간"] != detail_apps["최적사용시간"])
    ].copy()

    plans = data.plans.copy()
    settings_df = pd.DataFrame(list(data.settings.items()), columns=["설정항목", "값"])
    weights = data.weights.copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="요약_전체", index=False)
        detail_hourly.to_excel(writer, sheet_name="상세_시간대", index=False)
        detail_apps.to_excel(writer, sheet_name="상세_가전전체", index=False)
        changed.to_excel(writer, sheet_name="가전별_변경", index=False)
        plans.to_excel(writer, sheet_name="요금제", index=False)
        settings_df.to_excel(writer, sheet_name="사용설정", index=False)
        weights.to_excel(writer, sheet_name="시간대가중치", index=False)

    wb = load_workbook(output_path)
    dark = "17365D"
    blue = "4472C4"
    light = "D9EAF7"
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in list(col_cells)[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 34)

    # Formats and chart
    ws = wb["요약_전체"]
    headers = {cell.value: cell.column for cell in ws[1]}
    for name in ["구독료(원)", "구독최종납부액(원)", "현행요금_기준(원)", "현행요금_최적(원)"]:
        if name in headers:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, headers[name]).number_format = "#,##0"
    for name in [c for c in headers if isinstance(c, str) and ("kWh" in c or "kW)" in c or "100점" in c)]:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, headers[name]).number_format = "0.00"

    detail_ws = wb["상세_시간대"]
    detail_headers = {cell.value: cell.column for cell in detail_ws[1]}
    if "시간대" in detail_headers and "기준부하(kWh)" in detail_headers and "최적부하(kWh)" in detail_headers:
        chart = LineChart()
        chart.title = "제어 전·후 대표일 부하곡선"
        chart.y_axis.title = "시간당 전력량(kWh)"
        chart.x_axis.title = "시간대"
        data_cols = Reference(
            detail_ws,
            min_col=detail_headers["기준부하(kWh)"],
            max_col=detail_headers["최적부하(kWh)"],
            min_row=1,
            max_row=detail_ws.max_row,
        )
        cats = Reference(detail_ws, min_col=detail_headers["시간대"], min_row=2, max_row=detail_ws.max_row)
        chart.add_data(data_cols, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        detail_ws.add_chart(chart, "K2")

    # Notes
    notes = wb.create_sheet("읽어보기", 0)
    notes["A1"] = "구독형 전기요금 최적제어 결과 읽는 법"
    notes["A1"].fill = PatternFill("solid", fgColor=dark)
    notes["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    notes.merge_cells("A1:F1")
    note_lines = [
        "1. 요약_전체: 계절·요금제·냉난방모드·초과단가별 결과 비교",
        "2. 상세_시간대: 입력파일의 상세결과 설정에 해당하는 대표 시나리오의 주중·주말 부하곡선",
        "3. 가전별_변경: 실제로 사용시간이 이동하거나 사용량이 감축된 가전만 표시",
        "4. 고객편의지수는 상대 비교용 휴리스틱이며 실제 고객 만족도 측정값이 아님",
        "5. 현행요금 비교에서 기후환경요금·연료비조정요금은 입력파일에 기입한 단가를 사용",
        "6. 이 결과는 연구용 시뮬레이션이며 실제 설비 제어 전에는 통신·안전·고객동의 검증 필요",
    ]
    for i, line in enumerate(note_lines, 3):
        notes.cell(i, 1).value = line
    notes.column_dimensions["A"].width = 110
    notes.freeze_panes = "A3"
    wb.save(output_path)
    return output_path


def run_default_analysis(
    input_path: Union[str, os.PathLike],
    output_path: Union[str, os.PathLike] = "구독형_전기요금_최적화_결과.xlsx",
    customer_override: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    data = load_workbook_data(input_path)
    summary, hourly, apps = run_batch(data, customer_override=customer_override)
    output = export_results(output_path, summary, hourly, apps, data)
    return summary, hourly, apps, output


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="구독형 전기요금 최적제어 시뮬레이터")
    parser.add_argument("input", help="입력 Excel 파일")
    parser.add_argument("--output", default="구독형_전기요금_최적화_결과.xlsx")
    parser.add_argument("--override", action="store_true", help="고객 수동해제 적용: 자동제어 중단")
    args = parser.parse_args()
    summary, _, _, output = run_default_analysis(args.input, args.output, args.override)
    print(summary.head(12).to_string(index=False))
    print(f"\n결과 파일: {output}")
